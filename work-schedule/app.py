from flask import Flask, render_template
from dotenv import load_dotenv
import requests
import os
import pandas as pd
import openpyxl

load_dotenv()

## DEFINE GLOBAL VARIABLES
SPREADSHEET_LINK = os.getenv("SPREADSHEET_LINK")
SPREADSHEET_LINK_XLSX = os.getenv("SPREADSHEET_LINK_XLSX")
SHEET_NAME = os.getenv("SHEET_NAME")


app = Flask(__name__)


@app.route("/")
def index():
    """
    Fetches data from a specified spreadsheet link, processes it, and renders an HTML table.
    This function performs the following steps:
    1. Sends a GET request to fetch the spreadsheet from the provided link.
    2. Saves the spreadsheet content to a local file named 'spreadsheet.xlsx'.
    3. Loads the spreadsheet using the openpyxl library.
    4. Checks if the specified sheet name exists in the workbook.
    5. Extracts data from the sheet and processes it into a pandas DataFrame.
    6. Converts the DataFrame into an HTML table.
    7. Renders the HTML table in the 'index.html' template.
    Returns:
        str: Rendered HTML content for the 'index.html' template with the processed table data.
    """

    response = requests.get(SPREADSHEET_LINK_XLSX)

    with open("spreadsheet.xlsx", "wb") as f:
        f.write(response.content)

    workbook = openpyxl.load_workbook("spreadsheet.xlsx", data_only=True)
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        print(f"Sheet '{SHEET_NAME}' not found in the spreadsheet.")

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    headers = data[2]
    table_data = data[3:-1]

    headers = [header for header in headers if header]
    trimmed_table_data = []
    for row in table_data:
        trimmed_row = row[: len(headers)]
        trimmed_table_data.append(trimmed_row)

    df = pd.DataFrame(trimmed_table_data, columns=headers)
    df.fillna("", inplace=True)

    html_table = df.to_html(classes="styled-table", index=False)
    return render_template("index.html", table=html_table)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
